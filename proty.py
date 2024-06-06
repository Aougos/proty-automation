import argparse
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl_image_loader import SheetImageLoader
import os
from colorama import Fore, Style, init


print("-----------------------------------------------------------------------------------------------------------------------")
print("██████╗ ██████╗  ██████╗ ████████╗██╗   ██╗     █████╗ ██╗   ██╗████████╗ ██████╗ ███╗   ███╗ █████╗ ████████╗███████╗")
print("██╔══██╗██╔══██╗██╔═══██╗╚══██╔══╝╚██╗ ██╔╝    ██╔══██╗██║   ██║╚══██╔══╝██╔═══██╗████╗ ████║██╔══██╗╚══██╔══╝██╔════╝")
print("██████╔╝██████╔╝██║   ██║   ██║    ╚████╔╝     ███████║██║   ██║   ██║   ██║   ██║██╔████╔██║███████║   ██║   █████╗  ")
print("██╔═══╝ ██╔══██╗██║   ██║   ██║     ╚██╔╝      ██╔══██║██║   ██║   ██║   ██║   ██║██║╚██╔╝██║██╔══██║   ██║   ██╔══╝  ")
print("██║     ██║  ██║╚██████╔╝   ██║      ██║       ██║  ██║╚██████╔╝   ██║   ╚██████╔╝██║ ╚═╝ ██║██║  ██║   ██║   ███████╗")
print("╚═╝     ╚═╝  ╚═╝ ╚═════╝    ╚═╝      ╚═╝       ╚═╝  ╚═╝ ╚═════╝    ╚═╝    ╚═════╝ ╚═╝     ╚═╝╚═╝  ╚═╝   ╚═╝   ╚══════╝")
print("                                                 Work Happily")
print("-----------------------------------------------------------------------------------------------------------------------")


# Initialize colorama
init(autoreset=True)

def isImage(cell, image_loader):
    try:
        return image_loader.get(cell)
    except Exception as e:
        return None

def isNote(cell):
    if cell:
        return True
    else:
        return False

def count_testCase(file_path, sheet):
    threatModelling = pd.read_excel(file_path, sheet_name=sheet, skiprows=2)
    scenario_count = len(threatModelling)

    workbook = openpyxl.load_workbook(file_path)
    curr_sheet = workbook[sheet]
    image_loader = SheetImageLoader(curr_sheet)

    tempScenarioLoc   = ""
    tempScreenshotLoc = ""
    tempNoteLoc       = "" 
    for cell in curr_sheet[3]:  # The third row is accessed with curr_sheet[3]
        if tempScenarioLoc and tempScreenshotLoc and tempNoteLoc:
            break
        
        if not tempScreenshotLoc and cell.value == 'Screenshot Evidence':
            tempScreenshotLoc = get_column_letter(cell.column)
        
        if cell.value == 'Attack Scenario':
            tempScenarioLoc = get_column_letter(cell.column)
        
        if cell.value == 'Notes':
            tempNoteLoc = get_column_letter(cell.column)
            
    
    # print(f"Attack Scenario: {tempScenarioLoc}; Screen Shot: {tempScreenshotLoc}; Notes: {tempNoteLoc}")
    total_testcase = 0

    for x in range(4, scenario_count + 4):  # Ensure loop bounds are correct
        scenarioLoc     = tempScenarioLoc + str(x)
        screenshotLoc   = tempScreenshotLoc + str(x)
        noteLoc         = tempNoteLoc + str(x)  # Combine column letter and row number

        cell_screenshot = curr_sheet[screenshotLoc]
        cell_note = curr_sheet[noteLoc]

        is_image = isImage(screenshotLoc, image_loader)
        is_note = isNote(cell_note.value)

        screenshotNotNull = curr_sheet[screenshotLoc].value
        # print(screenshotNotNull, screenshotLoc)

        if curr_sheet[scenarioLoc].value is not None:
            if (is_image or screenshotNotNull is not None) or (is_image is None and is_note):
                total_testcase += 1
    
    print(f"    {Fore.RED}[!]{Style.RESET_ALL} Total Test Case of '{Fore.YELLOW}{sheet}{Style.RESET_ALL}' is: {Fore.YELLOW}{total_testcase}{Style.RESET_ALL} {Fore.GREEN}...done{Style.RESET_ALL}")
    return total_testcase

def count_sheet(file_path, filename):
    try:
        workbook = openpyxl.load_workbook(file_path)
        totalSheets = len(workbook.sheetnames)
        sheetNames = workbook.sheetnames
        clear_sheetNames = [sheet for sheet in sheetNames if sheet != 'Positive Testing']

        print(f"[*] Checking '{Fore.GREEN}Positive Testing{Style.RESET_ALL}' Sheet is exist {Fore.GREEN}...done{Style.RESET_ALL}")
        if 'Positive Testing' in sheetNames:
            totalTestcase = 0
            try:
                print(f"[*] Sheet '{Fore.GREEN}Positive Testing{Style.RESET_ALL}' is exists {Fore.GREEN}...done{Style.RESET_ALL}")
                print("[*] Calculating API...")
                apis = pd.read_excel(file_path, sheet_name='Positive Testing', skiprows=2)
                apis_count = apis['Feature'].count()
                loc = apis['Feature']
                print(f"[*] The number of APIs tested are: {Fore.YELLOW}{apis_count}{Style.RESET_ALL} {Fore.GREEN}...done{Style.RESET_ALL}")
                
                for sheet in clear_sheetNames:
                    tempSumTestCase = count_testCase(file_path, sheet)
                    totalTestcase += tempSumTestCase
                print(f"\n[*] Total Test Case for {Fore.YELLOW}{filename}{Style.RESET_ALL} are: {Fore.YELLOW}{totalTestcase}{Style.RESET_ALL}")
                print(f"{Fore.GREEN}[*] Proty has counted, now Proty wants to sleep{Style.RESET_ALL}")
                print(f"{Fore.GREEN}Proty sleep...{Style.RESET_ALL}")
                
            except Exception as e:
                print(f"An error occurred: {e}")

        else:
            print(f"{Fore.RED}[!] Sheet 'Positive Testing' does not exist. Cannot process counting Test Case{Style.RESET_ALL}")
            print(f"{Fore.RED}[!] Test Case Calculations Failed. Please Provide Correct Files{Style.RESET_ALL}")

    except FileNotFoundError:
        print(f"{Fore.RED}[!] File '{file_path}' not found.{Style.RESET_ALL}")
    except Exception as e:
        print(f"{Fore.RED}[!] An error occurred: {e}{Style.RESET_ALL}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Count test cases in an Excel file.")
    parser.add_argument('-f', '--file', type=str, help='Name of the Excel file without extension')
    parser.add_argument('-p', '--path', type=str, help='Path to the Excel file or directory containing the Excel file')

    args = parser.parse_args()

    if args.file:
        if args.path:
            file_path = os.path.join(args.path, f"{args.file}.xlsx")
        else:
            file_path = f"{args.file}.xlsx"
    elif args.path:
        file_path = args.path
    else:
        print(f"{Fore.RED}[!] You must provide either --file or --path or both.{Style.RESET_ALL}")
        exit(1)

    filename = os.path.basename(file_path)
    count_sheet(file_path, filename)